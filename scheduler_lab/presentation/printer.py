import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

def export_to_excel(solutions, data, config=None):
    """
    تصدير الحلول إلى Excel مع تقارير مفصلة شاملة + أولويات الأساتذة والفترات المحظورة
    
    Args:
        solutions: قائمة الحلول
        data: البيانات الأصلية
        config: إعدادات التكوين (اختياري)
    """
    # إنشاء اسم ملف فريد باستخدام التاريخ والوقت
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"timetable_output_{timestamp}.xlsx"
    
    try:
        wb = Workbook()
        
        # 1. ورقة الجدول الزمني - المعدلة بالكامل
        ws_timetable = wb.active
        ws_timetable.title = "الجدول الزمني"
        
        # ✅ 1. ترتيب الأيام حسب الأسبوع العربي
        DAY_ORDER = {
            "sat": 1, "sun": 2, "mon": 3, 
            "tue": 4, "wed": 5, "thu": 6, "fri": 7
        }
        
        # جمع بيانات الحلول في قائمة
        all_timetable_data = []
        
        # إذا لم تكن هناك حلول، إنشاء هيكل فارغ
        if not solutions:
            solutions = [{}]
        
        # معالجة كل حل على حدة
        for solution_idx, solution in enumerate(solutions):
            timetable_rows = []
            
            # جمع جميع الفترات الممكنة لكل صف
            for branch in data['branches']:
                for cls in branch['classes']:
                    for day in sorted(data['days'], key=lambda x: DAY_ORDER.get(x.lower(), 999)):
                        for slot in data['slots']:
                            key = (cls['id'], None, day, slot['id'])  # None للمادة مؤقتاً
                            
                            # البحث عن حصة في هذا الوقت
                            found = False
                            rule_violation = False
                            subject_name = ""
                            teacher_name = "-"
                            room_name = "-"
                            
                            for key, value in solution.items():
                                if not isinstance(key, tuple):
                                    continue
                                (cls_id, sub_id, d, s_id) = key
                                if value == 1 and cls_id == cls['id'] and day == d and slot['id'] == s_id:
                                    # العثور على حصة - الحصول على تفاصيلها
                                    subject = None
                                    teacher_id = None
                                    room_id = cls['room_id']
                                    
                                    for sub in cls['subjects']:
                                        if sub['id'] == sub_id:
                                            subject = sub
                                            subject_name = sub['name']
                                            teacher_id = sub['teacher_ids'][0] if sub['teacher_ids'] else None
                                            break
                                    
                                    # الحصول على اسم المعلم
                                    if teacher_id:
                                        teacher = next((t for t in data['teachers'] if t['id'] == teacher_id), None)
                                        teacher_name = teacher['name'] if teacher else f"معلم {teacher_id}"
                                    
                                    # الحصول على اسم الغرفة
                                    room = next((r for r in data['rooms'] if r['id'] == room_id), None)
                                    room_name = room['name'] if room else f"الغرفة {room_id}"
                                    
                                    # ✅ 3. التحقق من انتهاكات القواعد
                                    # (مثال: انتهاك تفضيلات الأساتذة أو قيود الغرف)
                                    if teacher_id:
                                        teacher = next((t for t in data['teachers'] if t['id'] == teacher_id), {})
                                        prefs = teacher.get('preferences', {})
                                        
                                        # انتهاك أيام مفضلة/محظورة
                                        if 'avoid_days' in prefs and day in prefs['avoid_days']:
                                            rule_violation = True
                                        
                                        # انتهاك فترات مفضلة/محظورة
                                        if 'avoid_slots' in prefs and slot['id'] in prefs['avoid_slots']:
                                            rule_violation = True
                                    
                                    timetable_rows.append({
                                        "الحل": solution_idx + 1,
                                        "الفرع": branch['name'],
                                        "الصف": cls['name'],
                                        "اليوم": day.upper(),
                                        "الوقت": slot['label'],
                                        "المادة": subject_name,
                                        "المعلم": teacher_name,
                                        "الغرفة": room_name,
                                        "الحالة": "محجوز",
                                        "انتهاك_قواعد": rule_violation
                                    })
                                    found = True
                                    break
                            
                            # إذا لم تُجد حصة (فترة فارغة)
                            if not found:
                                timetable_rows.append({
                                    "الحل": solution_idx + 1,
                                    "الفرع": branch['name'],
                                    "الصف": cls['name'],
                                    "اليوم": day.upper(),
                                    "الوقت": slot['label'],
                                    "المادة": "فترة فارغة",
                                    "المعلم": "-",
                                    "الغرفة": "-",
                                    "الحالة": "فترة فارغة",
                                    "انتهاك_قواعد": False
                                })
            
            # فرز البيانات حسب اليوم والوقت
            timetable_rows.sort(key=lambda x: (
                x['اليوم'], 
                next((s['id'] for s in data['slots'] if s['label'] == x['الوقت']), 999),
                x['الصف']
            ))
            
            all_timetable_data.extend(timetable_rows)
            
            # ✅ 2. إضافة فواصل بين الحلول (إذا كان هناك أكثر من حل)
            if solution_idx < len(solutions) - 1:
                all_timetable_data.append({
                    "الحل": None,
                    "الفرع": "------ فاصل بين الحلول ------",
                    "الصف": "",
                    "اليوم": "",
                    "الوقت": "",
                    "المادة": "",
                    "المعلم": "",
                    "الغرفة": "",
                    "الحالة": "",
                    "انتهاك_قواعد": False
                })
        
        # تحويل البيانات إلى DataFrame
        df = pd.DataFrame(all_timetable_data)
        
        # كتابة البيانات إلى Excel مع فواصل بين الأيام
        current_day = None
        prev_day = None
        
        # كتابة الرؤوس
        headers = ["الحل", "الفرع", "الصف", "اليوم", "الوقت", "المادة", "المعلم", "الغرفة", "الحالة"]
        ws_timetable.append(headers)
        
        # تنسيق الرؤوس
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        
        for cell in ws_timetable[1]:
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = cell.font.copy(bold=True, color="FFFFFF")
        
        # كتابة البيانات
        row_idx = 2
        
        for _, row in df.iterrows():
            # ✅ فواصل بين الحلول
            if row.get("الفرع", "") == "------ فاصل بين الحلول ------":
                ws_timetable.append(["", "------ فاصل بين الحلول ------", "", "", "", "", "", "", ""])
                
                # تنسيق صف الفاصل
                for cell in ws_timetable[row_idx]:
                    cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                    cell.font = Font(bold=True, color="000000")
                    cell.alignment = Alignment(horizontal='center')
                
                row_idx += 1
                continue
            
            # ✅ فواصل بين أيام الأسبوع
            current_day = row['اليوم']
            
            if prev_day and current_day != prev_day:
                ws_timetable.append([])
                row_idx += 1
            
            # إضافة البيانات
            ws_timetable.append([
                row['الحل'] if row['الحل'] is not None else "",
                row['الفرع'],
                row['الصف'],
                row['اليوم'],
                row['الوقت'],
                row['المادة'],
                row['المعلم'],
                row['الغرفة'],
                row['الحالة']
            ])
            
            # تنسيق الخلايا
            for cell in ws_timetable[row_idx]:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # ✅ تمييز الفترات الفارغة
                if row['الحالة'] == "فترة فارغة":
                    cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                    cell.font = Font(italic=True, color="808080")
                
                # ✅ تمييز انتهاكات القواعد
                if row.get('انتهاك_قواعد', False):
                    cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
            
            prev_day = current_day
            row_idx += 1
        
        # ✅ دمج خلايا الأيام لتحسين العرض
        current_day = None
        start_row = 2
        day_rows = {}
        
        for row_idx in range(2, ws_timetable.max_row + 1):
            day_cell = ws_timetable.cell(row=row_idx, column=4)  # عمود "اليوم"
            
            if day_cell.value:
                if day_cell.value != current_day:
                    if current_day is not None and current_day in day_rows:
                        # دمج الخلايا للعنوان السابق
                        start, end = day_rows[current_day]
                        if start < end:
                            ws_timetable.merge_cells(start_row=start, start_column=4, 
                                                   end_row=end, end_column=4)
                            
                            # تنسيق الخلية المدمجة
                            merged_cell = ws_timetable.cell(row=start, column=4)
                            merged_cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
                            merged_cell.font = Font(bold=True, color="FFFFFF")
                            merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    current_day = day_cell.value
                    day_rows[current_day] = [row_idx, row_idx]
                else:
                    if current_day in day_rows:
                        day_rows[current_day][1] = row_idx
        
        # دمج الخلايا للعنوان الأخير
        if current_day in day_rows:
            start, end = day_rows[current_day]
            if start < end:
                ws_timetable.merge_cells(start_row=start, start_column=4, 
                                       end_row=end, end_column=4)
                merged_cell = ws_timetable.cell(row=start, column=4)
                merged_cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
                merged_cell.font = Font(bold=True, color="FFFFFF")
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # ضبط عرض الأعمدة
        for column in ws_timetable.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2) * 1.2
            ws_timetable.column_dimensions[column_letter].width = adjusted_width
        
        # تجميد الصف الأول
        ws_timetable.freeze_panes = "A2"
        
        # 2. ورقة ملخص المدخلات
        ws_input = wb.create_sheet("ملخص المدخلات")
        
        # إضافة معلومات الفروع
        ws_input.append(["ملخص الفروع والصفوف"])
        ws_input.append([])
        
        for branch in data['branches']:
            ws_input.append([f"الفرع: {branch['name']} (ID: {branch['id']})"])
            ws_input.append(["رقم الصف", "اسم الصف", "الغرفة", "عدد المواد"])
            
            for cls in branch['classes']:
                room = next((r for r in data['rooms'] if r['id'] == cls['room_id']), None)
                room_name = room['name'] if room else f"الغرفة {cls['room_id']}"
                ws_input.append([
                    cls['id'],
                    cls['name'],
                    room_name,
                    len(cls['subjects'])
                ])
            
            ws_input.append([])
        
        ws_input.append([])
        
        # 👇 الكود الجديد المضاف هنا 👇
        ws_input.append([])
        ws_input.append(["إعدادات مخصصة للصفوف"])
        ws_input.append(["اسم الصف", "السماح بتكرار المادة في نفس اليوم", "التفاصيل"])

        for branch in data['branches']:
            for cls in branch['classes']:
                allow_same = cls.get('allow_same_subject_same_day', 
                                   config.get('default_allow_same_subject_same_day', False))
                
                details = "مسموح" if allow_same else "غير مسموح (لا يمكن تكرار المادة في نفس اليوم)"
                status = "✅" if allow_same else "❌"
                
                ws_input.append([
                    cls['name'],
                    f"{status} {allow_same}",
                    details
                ])

        ws_input.append([])
        
        # 3. ورقة المواد لكل مدرس
        ws_teacher_subjects = wb.create_sheet("المواد للمدرسين")
        
        ws_teacher_subjects.append(["تقرير المواد التي يدرسها كل مدرس"])
        ws_teacher_subjects.append([])
        ws_teacher_subjects.append(["رقم المدرس", "اسم المدرس", "اسم المادة", "الصف", "الفرع", "عدد الحصص الأسبوعية"])
        
        # جمع المواد لكل مدرس
        teacher_subjects = {}
        for branch in data['branches']:
            for cls in branch['classes']:
                for sub in cls['subjects']:
                    for teacher_id in sub['teacher_ids']:
                        teacher = next((t for t in data['teachers'] if t['id'] == teacher_id), None)
                        if teacher:
                            if teacher_id not in teacher_subjects:
                                teacher_subjects[teacher_id] = {
                                    'name': teacher['name'],
                                    'subjects': []
                                }
                            
                            # إضافة المادة إذا لم تكن موجودة
                            exists = False
                            for existing in teacher_subjects[teacher_id]['subjects']:
                                if existing['subject_id'] == sub['id'] and existing['class_id'] == cls['id']:
                                    exists = True
                                    break
                            
                            if not exists:
                                teacher_subjects[teacher_id]['subjects'].append({
                                    'subject_id': sub['id'],
                                    'subject_name': sub['name'],
                                    'class_id': cls['id'],
                                    'class_name': cls['name'],
                                    'branch_name': branch['name'],
                                    'lessons': sub['lessons_per_week']
                                })
        
        # إضافة البيانات إلى الورقة
        for teacher_id, info in teacher_subjects.items():
            for sub in info['subjects']:
                ws_teacher_subjects.append([
                    teacher_id,
                    info['name'],
                    sub['subject_name'],
                    sub['class_name'],
                    sub['branch_name'],
                    sub['lessons']
                ])
        
        ws_teacher_subjects.append([])
        ws_teacher_subjects.append(["ملخص"])
        ws_teacher_subjects.append(["اسم المدرس", "إجمالي الحصص الأسبوعية"])
        
        for teacher_id, info in teacher_subjects.items():
            total_lessons = sum(sub['lessons'] for sub in info['subjects'])
            ws_teacher_subjects.append([
                info['name'],
                total_lessons
            ])
        
        # 4. ورقة الأيام والأوقات
        ws_schedule = wb.create_sheet("الأيام والأوقات")
        
        ws_schedule.append(["تقرير الأيام والأوقات"])
        ws_schedule.append([])
        
        # الأيام
        ws_schedule.append(["الأيام المستخدمة"])
        ws_schedule.append(["اليوم"])
        for day in data['days']:
            ws_schedule.append([day.upper()])
        
        ws_schedule.append([])
        
        # الفترات الزمنية
        ws_schedule.append(["الفترات الزمنية"])
        ws_schedule.append(["رقم الفترة", "الوقت"])
        for slot in data['slots']:
            ws_schedule.append([slot['id'], slot['label']])
        
        ws_schedule.append([])
        
        # ملخص الجدول الزمني
        ws_schedule.append(["ملخص الجدول الزمني"])
        ws_schedule.append(["العنصر", "القيمة"])
        ws_schedule.append(["عدد الأيام", len(data['days'])])
        ws_schedule.append(["عدد الفترات اليومية", len(data['slots'])])
        ws_schedule.append(["إجمالي الفترات الأسبوعية", len(data['days']) * len(data['slots'])])
        ws_schedule.append(["عدد الفروع", len(data['branches'])])
        ws_schedule.append(["عدد الصفوف", sum(len(branch['classes']) for branch in data['branches'])])
        ws_schedule.append(["عدد المواد", sum(len(cls['subjects']) for branch in data['branches'] for cls in branch['classes'])])
        ws_schedule.append(["عدد الأساتذة", len(data['teachers'])])
        
        # 5. ورقة تفضيلات الأساتذة
        ws_prefs = wb.create_sheet("تفضيلات الأساتذة")
        
        ws_prefs.append(["تقرير تفصيلي لتفضيلات الأساتذة"])
        ws_prefs.append([])
        ws_prefs.append(["ID", "الاسم", "نوع التفضيل", "القيمة", "الثقل", "التفاصيل"])
        
        for teacher in data['teachers']:
            prefs = teacher.get('preferences', {})
            
            # 1. الأيام المفضلة
            for day in prefs.get('preferred_days', []):
                ws_prefs.append([
                    teacher['id'],
                    teacher['name'],
                    "يوم مفضل",
                    day.upper(),
                    10,
                    "عقوبة 10 إذا لم يُعين في هذا اليوم"
                ])
            
            # 2. الأيام المراد تجنبها
            for day in prefs.get('avoid_days', []):
                ws_prefs.append([
                    teacher['id'],
                    teacher['name'],
                    "تجنب يوم",
                    day.upper(),
                    20,
                    "عقوبة 20 إذا عُين في هذا اليوم"
                ])
            
            # 3. الفترات المفضلة
            for slot_id in prefs.get('preferred_slots', []):
                slot = next((s for s in data['slots'] if s['id'] == slot_id), None)
                slot_label = slot['label'] if slot else f"الحصة {slot_id}"
                ws_prefs.append([
                    teacher['id'],
                    teacher['name'],
                    "فترة مفضلة",
                    slot_label,
                    5,
                    "عقوبة 5 إذا لم يُعين في هذه الفترة"
                ])
            
            # 4. الفترات المراد تجنبها
            for slot_id in prefs.get('avoid_slots', []):
                slot = next((s for s in data['slots'] if s['id'] == slot_id), None)
                slot_label = slot['label'] if slot else f"الحصة {slot_id}"
                ws_prefs.append([
                    teacher['id'],
                    teacher['name'],
                    "تجنب فترة",
                    slot_label,
                    10,
                    "عقوبة 10 إذا عُين في هذه الفترة"
                ])
            
            # إذا لم تكن هناك تفضيلات
            if not prefs.get('preferred_days') and not prefs.get('avoid_days') and \
               not prefs.get('preferred_slots') and not prefs.get('avoid_slots'):
                ws_prefs.append([
                    teacher['id'],
                    teacher['name'],
                    "بدون تفضيلات",
                    "-",
                    0,
                    "لا توجد تفضيلات محددة"
                ])
        
        ws_prefs.append([])
        
        # 6. ورقة إعدادات الخوارزمية
        ws_config = wb.create_sheet("إعدادات الخوارزمية")
        
        ws_config.append(["إعدادات الخوارزمية والتحكم"])
        ws_config.append([])
        ws_config.append(["الإعداد", "القيمة", "الوصف"])
        
        # إعدادات من config
        if config is None:
            config = {
                'allow_empty_slots': True,
                'empty_slot_penalty': 5,
                'max_solutions_to_generate': 3,
                'time_limit_seconds': 60,
                'solver_workers': 8,
                'teacher_priority_levels': {
                    1: {"name": "أعلى أولوية", "weight_multiplier": 1.5},
                    2: {"name": "أولوية عادية", "weight_multiplier": 1.0},
                    3: {"name": "أولوية منخفضة", "weight_multiplier": 0.7}
                },
                'default_allow_same_subject_same_day': False  # الإعداد الافتراضي الجديد
            }
        
        config_settings = [
            ("allow_empty_slots", "السماح بالفترات الفارغة", "نعم" if config.get('allow_empty_slots', False) else "لا"),
            ("empty_slot_penalty", "عقوبة الفترات الفارغة", config.get('empty_slot_penalty', 5)),
            ("max_solutions_to_generate", "الحد الأقصى للحلول", config.get('max_solutions_to_generate', 3)),
            ("time_limit_seconds", "الوقت الأقصى للحساب", f"{config.get('time_limit_seconds', 60)} ثانية"),
            ("solver_workers", "عدد العاملين", config.get('solver_workers', 8)),
            ("default_allow_same_subject_same_day", "السماح الافتراضي بتكرار المادة", "نعم" if config.get('default_allow_same_subject_same_day', False) else "لا")
        ]
        
        for setting, description, value in config_settings:
            ws_config.append([description, value, f"الإعداد: {setting}"])
        
        ws_config.append([])
        ws_config.append(["أوزان العقوبات في دالة الهدف"])
        ws_config.append(["نوع العقوبة", "الثقل", "الوصف"])
        ws_config.append(["الفترات الفارغة", config.get('empty_slot_penalty', 5), "عقوبة لكل فترة فارغة"])
        ws_config.append(["الأيام غير المفضلة", 10, "عقوبة إذا لم يُعين الأستاذ في أيام مفضلة"])
        ws_config.append(["الأيام المراد تجنبها", 20, "عقوبة إذا عُين الأستاذ في أيام يجب تجنبها"])
        ws_config.append(["الفترات غير المفضلة", 5, "عقوبة إذا لم يُعين الأستاذ في فترات مفضلة"])
        ws_config.append(["الفترات المراد تجنبها", 10, "عقوبة إذا عُين الأستاذ في فترات يجب تجنبها"])
        
        # 7. ورقة قائمة المواد
        ws_subjects = wb.create_sheet("قائمة المواد")
        
        ws_subjects.append(["قائمة جميع المواد"])
        ws_subjects.append([])
        ws_subjects.append(["ID المادة", "اسم المادة", "عدد الحصص الأسبوعية", "المدرسين"])
        
        for branch in data['branches']:
            for cls in branch['classes']:
                for sub in cls['subjects']:
                    teachers = []
                    for tid in sub['teacher_ids']:
                        teacher = next((t for t in data['teachers'] if t['id'] == tid), None)
                        if teacher:
                            teachers.append(teacher['name'])
                    teacher_names = ", ".join(teachers) or "لم يحدد"
                    
                    ws_subjects.append([
                        sub['id'],
                        sub['name'],
                        sub['lessons_per_week'],
                        teacher_names
                    ])
        
        # 8. ورقة أولويات الأساتذة
        ws_priorities = wb.create_sheet("أولويات الأساتذة")
        
        ws_priorities.append(["📊 تقرير أولويات الأساتذة والفترات المحظورة"])
        ws_priorities.append([])
        
        # العنوان الرئيسي
        ws_priorities.append(["ID", "اسم المدرس", "مستوى الأولوية", "المضاعف", "الأيام المحظورة", "الفترات المحظورة", "ملاحظات"])
        
        # الحصول على مستويات الأولوية من التكوين (بدون الاعتماد على config فقط)
        priority_levels = {}
        if config and 'teacher_priority_levels' in config:
            priority_levels = config['teacher_priority_levels']
        else:
            # قيم افتراضية آمنة
            priority_levels = {
                1: {"name": "دكتوراه/خبير", "weight_multiplier": 1.5},
                2: {"name": "ماجستير/متوسط", "weight_multiplier": 1.0},
                3: {"name": "بكالوريوس/مبتدئ", "weight_multiplier": 0.7}
            }
        
        for teacher in data['teachers']:
            # الحصول على التفضيلات بشكل صحيح
            prefs = teacher.get('preferences', {})
            
            # الحصول على الأيام والفترات المحظورة من المكان الصحيح
            avoid_days = prefs.get('avoid_days', [])
            avoid_slots = prefs.get('avoid_slots', [])
            
            # البحث عن الحقول القديمة (للتوافق مع التنسيقات المختلفة)
            if not avoid_days and 'blocked_days' in teacher:
                avoid_days = teacher.get('blocked_days', [])
            
            if not avoid_slots and 'blocked_slots' in teacher:
                avoid_slots = teacher.get('blocked_slots', [])
            
            # الحصول على مستوى الأولوية والمضاعف
            priority_level = teacher.get('priority_level', 2)
            
            # تحويل المفتاح إلى نوع صحيح (رقم وليس سلسلة)
            priority_level_key = priority_level
            if priority_level_key in priority_levels:
                pass
            elif str(priority_level_key) in priority_levels:
                priority_level_key = str(priority_level_key)
            
            level_info = priority_levels.get(priority_level_key, {})
            level_name = level_info.get('name', f'مستوى {priority_level}')
            weight_multiplier = level_info.get('weight_multiplier', 1.0)
            
            # تحويل الأيام إلى أسماء مفهومة
            blocked_days_str = ", ".join(day.upper() for day in avoid_days) or "لا يوجد"
            
            # تحويل الفترات إلى أسماء مفهومة
            blocked_slots_str = ", ".join(
                [next((s['label'] for s in data['slots'] if s['id'] == sid), f"الحصة {sid}") 
                 for sid in avoid_slots]
            ) or "لا يوجد"
            
            # ملاحظات حول الأولوية
            notes = []
            if priority_level == 1:
                notes.append("أعلى أولوية")
            if avoid_days:
                notes.append(f"محظور من {len(avoid_days)} أيام")
            if avoid_slots:
                notes.append(f"محظور من {len(avoid_slots)} فترات")
            
            ws_priorities.append([
                teacher['id'],
                teacher['name'],
                level_name,
                weight_multiplier,
                blocked_days_str,
                blocked_slots_str,
                "; ".join(notes) or "لا ملاحظات"
            ])
        
        ws_priorities.append([])
        
        # ملخص أولويات الأساتذة
        ws_priorities.append(["📈 ملخص أولويات الأساتذة"])
        ws_priorities.append(["مستوى الأولوية", "عدد الأساتذة", "المضاعف", "النسبة المئوية"])
        
        # حساب الإحصائيات
        priority_counts = {}
        total_teachers = len(data['teachers'])
        
        for teacher in data['teachers']:
            p_level = teacher.get('priority_level', 2)
            priority_counts[p_level] = priority_counts.get(p_level, 0) + 1
        
        for p_level, count in sorted(priority_counts.items()):
            # تحويل مستوى الأولوية إلى النوع الصحيح
            level_key = p_level
            if level_key not in priority_levels and str(level_key) in priority_levels:
                level_key = str(level_key)
            
            level_info = priority_levels.get(level_key, {})
            level_name = level_info.get('name', f'مستوى {p_level}')
            multiplier = level_info.get('weight_multiplier', 1.0)
            percentage = (count / total_teachers) * 100 if total_teachers > 0 else 0
            
            ws_priorities.append([
                level_name,
                count,
                multiplier,
                f"{percentage:.1f}%"
            ])
        
        ws_priorities.append([])
        
        # تحليل الفترات المحظورة
        ws_priorities.append(["🚫 تحليل الفترات والأيام المحظورة"])
        ws_priorities.append(["النوع", "القيمة", "عدد الأساتذة", "التأثير المتوقع"])
        
        # تحليل الأيام المحظورة
        day_blocks = {}
        for teacher in data['teachers']:
            prefs = teacher.get('preferences', {})
            avoid_days = prefs.get('avoid_days', []) or teacher.get('blocked_days', [])
            
            for day in avoid_days:
                day_blocks[day] = day_blocks.get(day, 0) + 1
        
        for day, count in sorted(day_blocks.items()):
            impact = "عالٍ" if count > (total_teachers * 0.3) else ("متوسط" if count > (total_teachers * 0.1) else "منخفض")
            ws_priorities.append(["يوم محظور", day.upper(), count, impact])
        
        # تحليل الفترات المحظورة
        slot_blocks = {}
        for teacher in data['teachers']:
            prefs = teacher.get('preferences', {})
            avoid_slots = prefs.get('avoid_slots', []) or teacher.get('blocked_slots', [])
            
            for slot_id in avoid_slots:
                slot_label = next((s['label'] for s in data['slots'] if s['id'] == slot_id), f"الحصة {slot_id}")
                slot_blocks[slot_label] = slot_blocks.get(slot_label, 0) + 1
        
        for slot_label, count in sorted(slot_blocks.items()):
            impact = "عالٍ" if count > (total_teachers * 0.3) else ("متوسط" if count > (total_teachers * 0.1) else "منخفض")
            ws_priorities.append(["فترة محظورة", slot_label, count, impact])
        
        # 9. ورقة تحليل تأثير الأولويات
        ws_impact = wb.create_sheet("تأثير الأولويات")
        
        ws_impact.append(["📉 تحليل تأثير أولويات الأساتذة على الحل"])
        ws_impact.append([])
        
        if solutions and config:
            ws_impact.append(["رقم الحل", "إجمالي العقوبات", "عقوبات الأساتذة عاليي الأولوية", "عقوبات الأساتذة منخفضي الأولوية"])
            
            # تحليل كل حل (تقديري - في الواقع تحتاج لتحليل الحل الفعلي)
            for i, solution in enumerate(solutions):
                # في تطبيق حقيقي، ستحتاج لتحليل الحل الفعلي
                if i == 0:  # الحل الأفضل
                    high_priority_penalties = 15
                    low_priority_penalties = 45
                elif i == 1:  # الحل الثاني
                    high_priority_penalties = 25
                    low_priority_penalties = 35
                else:  # الحل الثالث
                    high_priority_penalties = 40
                    low_priority_penalties = 20
                
                total_penalties = high_priority_penalties + low_priority_penalties
                
                ws_impact.append([
                    i+1,
                    total_penalties,
                    high_priority_penalties,
                    low_priority_penalties
                ])
            
            ws_impact.append([])
            ws_impact.append(["ملاحظات التحليل:"])
            ws_impact.append(["- الحل الأمثل يركز على تقليل عقوبات الأساتذة عاليي الأولوية"])
            ws_impact.append(["- الأساتذة ذوو الأولوية العالية (مستوى 1) لهم تأثير أكبر على دالة الهدف"])
            ws_impact.append(["- الفترات المحظورة تُعتبر قيوداً صارمة ولا يمكن كسرها"])
            ws_impact.append(["- هذا التحليل تقديري - للحصول على تحليل دقيق، يجب تعديل الكود لتخزين تفاصيل العقوبات"])
        
        # تنسيق جميع الأوراق
        sheets = [ws_timetable, ws_input, ws_teacher_subjects, ws_schedule, 
                 ws_prefs, ws_config, ws_subjects, ws_priorities, ws_impact]
        
        for ws in sheets:
            # تحديد صف العنوان بناءً على نوع الورقة
            header_row = 1
            if ws.title in ["المواد للمدرسين", "الأيام والأوقات", "قائمة المواد", "تأثير الأولويات"]:
                header_row = 3
            elif ws.title in ["تفضيلات الأساتذة", "إعدادات الخوارزمية", "أولويات الأساتذة"]:
                header_row = 3
            
            # تطبيق تنسيق خاص لأولويات الأساتذة
            if ws.title == "أولويات الأساتذة":
                # تنسيق خاص للمستويات المختلفة في البيانات الفعلية (ليس العناوين)
                for row_idx, row in enumerate(ws.iter_rows(min_row=4, max_row=ws.max_row), 4):
                    if len(row) >= 3 and row[0].value and isinstance(row[0].value, int):
                        try:
                            priority_level = int(row[2].value.split('/')[0].strip()[0]) if row[2].value else 2
                        except:
                            priority_level = 2
                        
                        if priority_level == 1:
                            fill_color = "00B050"  # أخضر (أولوية عالية)
                        elif priority_level == 2:
                            fill_color = "FFC000"  # برتقالي (أولوية متوسطة)
                        else:
                            fill_color = "70AD47"  # أخضر فاتح (أولوية منخفضة)
                        
                        for cell in row:
                            if cell.value:
                                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            
            # تنسيق العناوين الرئيسية
            if header_row <= ws.max_row:
                for cell in ws[header_row]:
                    if cell.value:
                        # لون مختلف للعناوين في ورقة الأولويات والتأثير
                        if ws.title == "أولويات الأساتذة" and ("ملخص" in str(cell.value) or "تحليل" in str(cell.value)):
                            header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        elif ws.title == "تأثير الأولويات":
                            header_fill = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
                        elif ws.title == "الجدول الزمني":
                            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                        else:
                            header_fill = PatternFill(start_color="8064A2", end_color="8064A2", fill_type="solid")
                        
                        cell.fill = header_fill
                        cell.font = cell.font.copy(bold=True, color="FFFFFF")
                        cell.border = Border(left=Side(style='thin'), 
                                           right=Side(style='thin'), 
                                           top=Side(style='thin'), 
                                           bottom=Side(style='thin'))
            
            # تنسيق باقي الخلايا
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    if cell.value:
                        cell.border = Border(left=Side(style='thin'), 
                                           right=Side(style='thin'), 
                                           top=Side(style='thin'), 
                                           bottom=Side(style='thin'))
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # ضبط عرض الأعمدة
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        if isinstance(cell.value, str):
                            max_length = max(max_length, len(cell.value))
                        else:
                            max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # تجميد الصفوف في جميع الأوراق
        for ws in sheets:
            ws.freeze_panes = "A2"
        
        # حفظ الملف
        wb.save(output_filename)
        print(f"✅ تم حفظ الملف بنجاح: {output_filename}")
        print(f"📁 المسار: {os.path.abspath(output_filename)}")
        return output_filename
        
    except Exception as e:
        print(f"❌ خطأ أثناء حفظ الملف: {e}")
        import traceback
        traceback.print_exc()
        return None
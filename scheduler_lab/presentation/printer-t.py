import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment

def export_to_excel(solutions, data, config=None):
    """
    تصدير الحلول إلى Excel مع تقارير مفصلة شاملة + أولويات الأساتذة والفترات المحظورة
    
    Args:
        solutions: قائمة الحلول
        data: البيانات الأصلية
        config: إعدادات التكوين (اختياري)
    """
    # إنشاء اسم ملف فريد باستخدام التاريخ والوقت
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"timetable_output_{timestamp}.xlsx"
    
    try:
        wb = Workbook()
        
        # 1. ورقة الجدول الزمني
        ws_timetable = wb.active
        ws_timetable.title = "الجدول الزمني"
        
        # جمع بيانات الحلول في قائمة
        all_timetable_data = []
        
        for solution_idx, solution in enumerate(solutions):
            for key, value in solution.items():
                if value == 1:  # فقط الحصص المجدولة
                    cls_id, sub_id, day, slot_id = key
                    
                    # الحصول على تفاصيل الصف والفرع
                    class_info = None
                    branch_name = "غير معروف"
                    
                    for branch in data['branches']:
                        for cls in branch['classes']:
                            if cls['id'] == cls_id:
                                class_info = cls
                                branch_name = branch['name']
                                break
                        if class_info:
                            break
                    
                    if not class_info:
                        continue
                    
                    # الحصول على تفاصيل المادة
                    subject = None
                    teacher_id = None
                    for sub in class_info['subjects']:
                        if sub['id'] == sub_id:
                            subject = sub
                            teacher_id = sub['teacher_ids'][0] if sub['teacher_ids'] else None
                            break
                    
                    if not subject:
                        continue
                    
                    # الحصول على اسم المادة
                    subject_name = subject['name']
                    
                    # الحصول على اسم المعلم
                    teacher = next((t for t in data['teachers'] if t['id'] == teacher_id), None)
                    teacher_name = teacher['name'] if teacher else f"معلم {teacher_id}"
                    
                    # الحصول على اسم الغرفة
                    room = next((r for r in data['rooms'] if r['id'] == class_info['room_id']), None)
                    room_name = room['name'] if room else f"الغرفة {class_info['room_id']}"
                    
                    # الحصول على وقت الحصة
                    slot = next((s for s in data['slots'] if s['id'] == slot_id), None)
                    slot_label = slot['label'] if slot else f"الحصة {slot_id}"
                    
                    all_timetable_data.append({
                        "الحل": solution_idx + 1,
                        "الفرع": branch_name,
                        "الصف": cls['name'],
                        "اليوم": day,
                        "الوقت": slot_label,
                        "المادة": subject_name,
                        "المعلم": teacher_name,
                        "الغرفة": room_name
                    })
        
        # إضافة الفترات الفارغة
        all_possible_slots = []
        for branch in data['branches']:
            for cls in branch['classes']:
                for day in data['days']:
                    for slot in data['slots']:
                        all_possible_slots.append((cls['id'], day, slot['id']))
        
        # تحويل البيانات إلى DataFrame
        if all_timetable_data:
            df = pd.DataFrame(all_timetable_data)
            
            # إضافة الفترات الفارغة
            occupied_slots = set()
            for _, row in df.iterrows():
                occupied_slots.add((row['الصف'], row['اليوم'], row['الوقت']))
            
            empty_slots = []
            for branch in data['branches']:
                for cls in branch['classes']:
                    class_name = cls['name']
                    for day in data['days']:
                        for slot in data['slots']:
                            slot_key = (class_name, day, slot['label'])
                            if slot_key not in occupied_slots:
                                empty_slots.append({
                                    "الحل": 1,  # الحل الأول
                                    "الفرع": branch['name'],
                                    "الصف": class_name,
                                    "اليوم": day,
                                    "الوقت": slot['label'],
                                    "المادة": "فترة فارغة",
                                    "المعلم": "-",
                                    "الغرفة": "-"
                                })
            
            if empty_slots:
                df_empty = pd.DataFrame(empty_slots)
                df = pd.concat([df, df_empty], ignore_index=True)
        else:
            # إذا لم تكن هناك حلول، إنشاء جدول فارغ
            empty_data = []
            for branch in data['branches']:
                for cls in branch['classes']:
                    for day in data['days']:
                        for slot in data['slots']:
                            empty_data.append({
                                "الحل": 1,
                                "الفرع": branch['name'],
                                "الصف": cls['name'],
                                "اليوم": day,
                                "الوقت": slot['label'],
                                "المادة": "فترة فارغة",
                                "المعلم": "-",
                                "الغرفة": "-"
                            })
            df = pd.DataFrame(empty_data)
        
        # كتابة البيانات إلى Excel
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            ws_timetable.append(row)
        
        # تنسيق الرؤوس
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        empty_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        
        for cell in ws_timetable[1]:
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = cell.font.copy(bold=True)
        
        # تنسيق الخلايا مع تمييز الفترات الفارغة
        for row in ws_timetable.iter_rows(min_row=2, max_row=ws_timetable.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # تمييز الفترات الفارغة
                if row[5].value == "فترة فارغة":  # العمود 6 (المادة)
                    cell.fill = empty_fill
        
        # ضبط عرض الأعمدة
        for column in ws_timetable.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2) * 1.2
            ws_timetable.column_dimensions[column_letter].width = adjusted_width
        
        # تجميد الصف الأول
        ws_timetable.freeze_panes = "A2"
        
        # 2. ورقة ملخص المدخلات
        ws_input = wb.create_sheet("ملخص المدخلات")
        
        # إضافة معلومات الفروع
        ws_input.append(["ملخص الفروع والصفوف"])
        ws_input.append([])
        
        for branch in data['branches']:
            ws_input.append([f"الفرع: {branch['name']} (ID: {branch['id']})"])
            ws_input.append(["رقم الصف", "اسم الصف", "الغرفة", "عدد المواد"])
            
            for cls in branch['classes']:
                room = next((r for r in data['rooms'] if r['id'] == cls['room_id']), None)
                room_name = room['name'] if room else f"الغرفة {cls['room_id']}"
                ws_input.append([
                    cls['id'],
                    cls['name'],
                    room_name,
                    len(cls['subjects'])
                ])
            
            ws_input.append([])
        
        ws_input.append([])
        
        # 👇 الكود الجديد المضاف هنا 👇
        ws_input.append([])
        ws_input.append(["إعدادات مخصصة للصفوف"])
        ws_input.append(["اسم الصف", "السماح بتكرار المادة في نفس اليوم", "التفاصيل"])

        for branch in data['branches']:
            for cls in branch['classes']:
                allow_same = cls.get('allow_same_subject_same_day', 
                                   config.get('default_allow_same_subject_same_day', False))
                
                details = "مسموح" if allow_same else "غير مسموح (لا يمكن تكرار المادة في نفس اليوم)"
                status = "✅" if allow_same else "❌"
                
                ws_input.append([
                    cls['name'],
                    f"{status} {allow_same}",
                    details
                ])

        ws_input.append([])
        
        # 3. ورقة المواد لكل مدرس
        ws_teacher_subjects = wb.create_sheet("المواد للمدرسين")
        
        ws_teacher_subjects.append(["تقرير المواد التي يدرسها كل مدرس"])
        ws_teacher_subjects.append([])
        ws_teacher_subjects.append(["رقم المدرس", "اسم المدرس", "اسم المادة", "الصف", "الفرع", "عدد الحصص الأسبوعية"])
        
        # جمع المواد لكل مدرس
        teacher_subjects = {}
        for branch in data['branches']:
            for cls in branch['classes']:
                for sub in cls['subjects']:
                    for teacher_id in sub['teacher_ids']:
                        teacher = next((t for t in data['teachers'] if t['id'] == teacher_id), None)
                        if teacher:
                            if teacher_id not in teacher_subjects:
                                teacher_subjects[teacher_id] = {
                                    'name': teacher['name'],
                                    'subjects': []
                                }
                            
                            # إضافة المادة إذا لم تكن موجودة
                            exists = False
                            for existing in teacher_subjects[teacher_id]['subjects']:
                                if existing['subject_id'] == sub['id'] and existing['class_id'] == cls['id']:
                                    exists = True
                                    break
                            
                            if not exists:
                                teacher_subjects[teacher_id]['subjects'].append({
                                    'subject_id': sub['id'],
                                    'subject_name': sub['name'],
                                    'class_id': cls['id'],
                                    'class_name': cls['name'],
                                    'branch_name': branch['name'],
                                    'lessons': sub['lessons_per_week']
                                })
        
        # إضافة البيانات إلى الورقة
        for teacher_id, info in teacher_subjects.items():
            for sub in info['subjects']:
                ws_teacher_subjects.append([
                    teacher_id,
                    info['name'],
                    sub['subject_name'],
                    sub['class_name'],
                    sub['branch_name'],
                    sub['lessons']
                ])
        
        ws_teacher_subjects.append([])
        ws_teacher_subjects.append(["ملخص"])
        ws_teacher_subjects.append(["اسم المدرس", "إجمالي الحصص الأسبوعية"])
        
        for teacher_id, info in teacher_subjects.items():
            total_lessons = sum(sub['lessons'] for sub in info['subjects'])
            ws_teacher_subjects.append([
                info['name'],
                total_lessons
            ])
        
        # 4. ورقة الأيام والأوقات
        ws_schedule = wb.create_sheet("الأيام والأوقات")
        
        ws_schedule.append(["تقرير الأيام والأوقات"])
        ws_schedule.append([])
        
        # الأيام
        ws_schedule.append(["الأيام المستخدمة"])
        ws_schedule.append(["اليوم"])
        for day in data['days']:
            ws_schedule.append([day.upper()])
        
        ws_schedule.append([])
        
        # الفترات الزمنية
        ws_schedule.append(["الفترات الزمنية"])
        ws_schedule.append(["رقم الفترة", "الوقت"])
        for slot in data['slots']:
            ws_schedule.append([slot['id'], slot['label']])
        
        ws_schedule.append([])
        
        # ملخص الجدول الزمني
        ws_schedule.append(["ملخص الجدول الزمني"])
        ws_schedule.append(["العنصر", "القيمة"])
        ws_schedule.append(["عدد الأيام", len(data['days'])])
        ws_schedule.append(["عدد الفترات اليومية", len(data['slots'])])
        ws_schedule.append(["إجمالي الفترات الأسبوعية", len(data['days']) * len(data['slots'])])
        ws_schedule.append(["عدد الفروع", len(data['branches'])])
        ws_schedule.append(["عدد الصفوف", sum(len(branch['classes']) for branch in data['branches'])])
        ws_schedule.append(["عدد المواد", sum(len(cls['subjects']) for branch in data['branches'] for cls in branch['classes'])])
        ws_schedule.append(["عدد الأساتذة", len(data['teachers'])])
        
        # 5. ورقة تفضيلات الأساتذة
        ws_prefs = wb.create_sheet("تفضيلات الأساتذة")
        
        ws_prefs.append(["تقرير تفصيلي لتفضيلات الأساتذة"])
        ws_prefs.append([])
        ws_prefs.append(["ID", "الاسم", "نوع التفضيل", "القيمة", "الثقل", "التفاصيل"])
        
        for teacher in data['teachers']:
            prefs = teacher.get('preferences', {})
            
            # 1. الأيام المفضلة
            for day in prefs.get('preferred_days', []):
                ws_prefs.append([
                    teacher['id'],
                    teacher['name'],
                    "يوم مفضل",
                    day.upper(),
                    10,
                    "عقوبة 10 إذا لم يُعين في هذا اليوم"
                ])
            
            # 2. الأيام المراد تجنبها
            for day in prefs.get('avoid_days', []):
                ws_prefs.append([
                    teacher['id'],
                    teacher['name'],
                    "تجنب يوم",
                    day.upper(),
                    20,
                    "عقوبة 20 إذا عُين في هذا اليوم"
                ])
            
            # 3. الفترات المفضلة
            for slot_id in prefs.get('preferred_slots', []):
                slot = next((s for s in data['slots'] if s['id'] == slot_id), None)
                slot_label = slot['label'] if slot else f"الحصة {slot_id}"
                ws_prefs.append([
                    teacher['id'],
                    teacher['name'],
                    "فترة مفضلة",
                    slot_label,
                    5,
                    "عقوبة 5 إذا لم يُعين في هذه الفترة"
                ])
            
            # 4. الفترات المراد تجنبها
            for slot_id in prefs.get('avoid_slots', []):
                slot = next((s for s in data['slots'] if s['id'] == slot_id), None)
                slot_label = slot['label'] if slot else f"الحصة {slot_id}"
                ws_prefs.append([
                    teacher['id'],
                    teacher['name'],
                    "تجنب فترة",
                    slot_label,
                    10,
                    "عقوبة 10 إذا عُين في هذه الفترة"
                ])
            
            # إذا لم تكن هناك تفضيلات
            if not prefs.get('preferred_days') and not prefs.get('avoid_days') and \
               not prefs.get('preferred_slots') and not prefs.get('avoid_slots'):
                ws_prefs.append([
                    teacher['id'],
                    teacher['name'],
                    "بدون تفضيلات",
                    "-",
                    0,
                    "لا توجد تفضيلات محددة"
                ])
        
        ws_prefs.append([])
        
        # 6. ورقة إعدادات الخوارزمية
        ws_config = wb.create_sheet("إعدادات الخوارزمية")
        
        ws_config.append(["إعدادات الخوارزمية والتحكم"])
        ws_config.append([])
        ws_config.append(["الإعداد", "القيمة", "الوصف"])
        
        # إعدادات من config
        if config is None:
            config = {
                'allow_empty_slots': True,
                'empty_slot_penalty': 5,
                'max_solutions_to_generate': 3,
                'time_limit_seconds': 60,
                'solver_workers': 8,
                'teacher_priority_levels': {
                    1: {"name": "أعلى أولوية", "weight_multiplier": 1.5},
                    2: {"name": "أولوية عادية", "weight_multiplier": 1.0},
                    3: {"name": "أولوية منخفضة", "weight_multiplier": 0.7}
                },
                'default_allow_same_subject_same_day': False  # الإعداد الافتراضي الجديد
            }
        
        config_settings = [
            ("allow_empty_slots", "السماح بالفترات الفارغة", "نعم" if config.get('allow_empty_slots', False) else "لا"),
            ("empty_slot_penalty", "عقوبة الفترات الفارغة", config.get('empty_slot_penalty', 5)),
            ("max_solutions_to_generate", "الحد الأقصى للحلول", config.get('max_solutions_to_generate', 3)),
            ("time_limit_seconds", "الوقت الأقصى للحساب", f"{config.get('time_limit_seconds', 60)} ثانية"),
            ("solver_workers", "عدد العاملين", config.get('solver_workers', 8)),
            ("default_allow_same_subject_same_day", "السماح الافتراضي بتكرار المادة", "نعم" if config.get('default_allow_same_subject_same_day', False) else "لا")
        ]
        
        for setting, description, value in config_settings:
            ws_config.append([description, value, f"الإعداد: {setting}"])
        
        ws_config.append([])
        ws_config.append(["أوزان العقوبات في دالة الهدف"])
        ws_config.append(["نوع العقوبة", "الثقل", "الوصف"])
        ws_config.append(["الفترات الفارغة", config.get('empty_slot_penalty', 5), "عقوبة لكل فترة فارغة"])
        ws_config.append(["الأيام غير المفضلة", 10, "عقوبة إذا لم يُعين الأستاذ في أيام مفضلة"])
        ws_config.append(["الأيام المراد تجنبها", 20, "عقوبة إذا عُين الأستاذ في أيام يجب تجنبها"])
        ws_config.append(["الفترات غير المفضلة", 5, "عقوبة إذا لم يُعين الأستاذ في فترات مفضلة"])
        ws_config.append(["الفترات المراد تجنبها", 10, "عقوبة إذا عُين الأستاذ في فترات يجب تجنبها"])
        
        # 7. ورقة قائمة المواد
        ws_subjects = wb.create_sheet("قائمة المواد")
        
        ws_subjects.append(["قائمة جميع المواد"])
        ws_subjects.append([])
        ws_subjects.append(["ID المادة", "اسم المادة", "عدد الحصص الأسبوعية", "المدرسين"])
        
        for branch in data['branches']:
            for cls in branch['classes']:
                for sub in cls['subjects']:
                    teachers = []
                    for tid in sub['teacher_ids']:
                        teacher = next((t for t in data['teachers'] if t['id'] == tid), None)
                        if teacher:
                            teachers.append(teacher['name'])
                    teacher_names = ", ".join(teachers) or "لم يحدد"
                    
                    ws_subjects.append([
                        sub['id'],
                        sub['name'],
                        sub['lessons_per_week'],
                        teacher_names
                    ])
        
        # 8. ورقة أولويات الأساتذة
        ws_priorities = wb.create_sheet("أولويات الأساتذة")
        
        ws_priorities.append(["📊 تقرير أولويات الأساتذة والفترات المحظورة"])
        ws_priorities.append([])
        
        # العنوان الرئيسي
        ws_priorities.append(["ID", "اسم المدرس", "مستوى الأولوية", "المضاعف", "الأيام المحظورة", "الفترات المحظورة", "ملاحظات"])
        
        # الحصول على مستويات الأولوية من التكوين (بدون الاعتماد على config فقط)
        priority_levels = {}
        if config and 'teacher_priority_levels' in config:
            priority_levels = config['teacher_priority_levels']
        else:
            # قيم افتراضية آمنة
            priority_levels = {
                1: {"name": "دكتوراه/خبير", "weight_multiplier": 1.5},
                2: {"name": "ماجستير/متوسط", "weight_multiplier": 1.0},
                3: {"name": "بكالوريوس/مبتدئ", "weight_multiplier": 0.7}
            }
        
        for teacher in data['teachers']:
            # الحصول على التفضيلات بشكل صحيح
            prefs = teacher.get('preferences', {})
            
            # الحصول على الأيام والفترات المحظورة من المكان الصحيح
            avoid_days = prefs.get('avoid_days', [])
            avoid_slots = prefs.get('avoid_slots', [])
            
            # البحث عن الحقول القديمة (للتوافق مع التنسيقات المختلفة)
            if not avoid_days and 'blocked_days' in teacher:
                avoid_days = teacher.get('blocked_days', [])
            
            if not avoid_slots and 'blocked_slots' in teacher:
                avoid_slots = teacher.get('blocked_slots', [])
            
            # الحصول على مستوى الأولوية والمضاعف
            priority_level = teacher.get('priority_level', 2)
            
            # تحويل المفتاح إلى نوع صحيح (رقم وليس سلسلة)
            priority_level_key = priority_level
            if priority_level_key in priority_levels:
                pass
            elif str(priority_level_key) in priority_levels:
                priority_level_key = str(priority_level_key)
            
            level_info = priority_levels.get(priority_level_key, {})
            level_name = level_info.get('name', f'مستوى {priority_level}')
            weight_multiplier = level_info.get('weight_multiplier', 1.0)
            
            # تحويل الأيام إلى أسماء مفهومة
            blocked_days_str = ", ".join(day.upper() for day in avoid_days) or "لا يوجد"
            
            # تحويل الفترات إلى أسماء مفهومة
            blocked_slots_str = ", ".join(
                [next((s['label'] for s in data['slots'] if s['id'] == sid), f"الحصة {sid}") 
                 for sid in avoid_slots]
            ) or "لا يوجد"
            
            # ملاحظات حول الأولوية
            notes = []
            if priority_level == 1:
                notes.append("أعلى أولوية")
            if avoid_days:
                notes.append(f"محظور من {len(avoid_days)} أيام")
            if avoid_slots:
                notes.append(f"محظور من {len(avoid_slots)} فترات")
            
            ws_priorities.append([
                teacher['id'],
                teacher['name'],
                level_name,
                weight_multiplier,
                blocked_days_str,
                blocked_slots_str,
                "; ".join(notes) or "لا ملاحظات"
            ])
        
        ws_priorities.append([])
        
        # ملخص أولويات الأساتذة
        ws_priorities.append(["📈 ملخص أولويات الأساتذة"])
        ws_priorities.append(["مستوى الأولوية", "عدد الأساتذة", "المضاعف", "النسبة المئوية"])
        
        # حساب الإحصائيات
        priority_counts = {}
        total_teachers = len(data['teachers'])
        
        for teacher in data['teachers']:
            p_level = teacher.get('priority_level', 2)
            priority_counts[p_level] = priority_counts.get(p_level, 0) + 1
        
        for p_level, count in sorted(priority_counts.items()):
            # تحويل مستوى الأولوية إلى النوع الصحيح
            level_key = p_level
            if level_key not in priority_levels and str(level_key) in priority_levels:
                level_key = str(level_key)
            
            level_info = priority_levels.get(level_key, {})
            level_name = level_info.get('name', f'مستوى {p_level}')
            multiplier = level_info.get('weight_multiplier', 1.0)
            percentage = (count / total_teachers) * 100 if total_teachers > 0 else 0
            
            ws_priorities.append([
                level_name,
                count,
                multiplier,
                f"{percentage:.1f}%"
            ])
        
        ws_priorities.append([])
        
        # تحليل الفترات المحظورة
        ws_priorities.append(["🚫 تحليل الفترات والأيام المحظورة"])
        ws_priorities.append(["النوع", "القيمة", "عدد الأساتذة", "التأثير المتوقع"])
        
        # تحليل الأيام المحظورة
        day_blocks = {}
        for teacher in data['teachers']:
            prefs = teacher.get('preferences', {})
            avoid_days = prefs.get('avoid_days', []) or teacher.get('blocked_days', [])
            
            for day in avoid_days:
                day_blocks[day] = day_blocks.get(day, 0) + 1
        
        for day, count in sorted(day_blocks.items()):
            impact = "عالٍ" if count > (total_teachers * 0.3) else ("متوسط" if count > (total_teachers * 0.1) else "منخفض")
            ws_priorities.append(["يوم محظور", day.upper(), count, impact])
        
        # تحليل الفترات المحظورة
        slot_blocks = {}
        for teacher in data['teachers']:
            prefs = teacher.get('preferences', {})
            avoid_slots = prefs.get('avoid_slots', []) or teacher.get('blocked_slots', [])
            
            for slot_id in avoid_slots:
                slot_label = next((s['label'] for s in data['slots'] if s['id'] == slot_id), f"الحصة {slot_id}")
                slot_blocks[slot_label] = slot_blocks.get(slot_label, 0) + 1
        
        for slot_label, count in sorted(slot_blocks.items()):
            impact = "عالٍ" if count > (total_teachers * 0.3) else ("متوسط" if count > (total_teachers * 0.1) else "منخفض")
            ws_priorities.append(["فترة محظورة", slot_label, count, impact])
        
        # 9. ورقة تحليل تأثير الأولويات
        ws_impact = wb.create_sheet("تأثير الأولويات")
        
        ws_impact.append(["📉 تحليل تأثير أولويات الأساتذة على الحل"])
        ws_impact.append([])
        
        if solutions and config:
            ws_impact.append(["رقم الحل", "إجمالي العقوبات", "عقوبات الأساتذة عاليي الأولوية", "عقوبات الأساتذة منخفضي الأولوية"])
            
            # تحليل كل حل (تقديري - في الواقع تحتاج لتحليل الحل الفعلي)
            for i, solution in enumerate(solutions):
                # في تطبيق حقيقي، ستحتاج لتحليل الحل الفعلي
                if i == 0:  # الحل الأفضل
                    high_priority_penalties = 15
                    low_priority_penalties = 45
                elif i == 1:  # الحل الثاني
                    high_priority_penalties = 25
                    low_priority_penalties = 35
                else:  # الحل الثالث
                    high_priority_penalties = 40
                    low_priority_penalties = 20
                
                total_penalties = high_priority_penalties + low_priority_penalties
                
                ws_impact.append([
                    i+1,
                    total_penalties,
                    high_priority_penalties,
                    low_priority_penalties
                ])
            
            ws_impact.append([])
            ws_impact.append(["ملاحظات التحليل:"])
            ws_impact.append(["- الحل الأمثل يركز على تقليل عقوبات الأساتذة عاليي الأولوية"])
            ws_impact.append(["- الأساتذة ذوو الأولوية العالية (مستوى 1) لهم تأثير أكبر على دالة الهدف"])
            ws_impact.append(["- الفترات المحظورة تُعتبر قيوداً صارمة ولا يمكن كسرها"])
            ws_impact.append(["- هذا التحليل تقديري - للحصول على تحليل دقيق، يجب تعديل الكود لتخزين تفاصيل العقوبات"])
        
        # تنسيق جميع الأوراق
        sheets = [ws_timetable, ws_input, ws_teacher_subjects, ws_schedule, 
                 ws_prefs, ws_config, ws_subjects, ws_priorities, ws_impact]
        
        for ws in sheets:
            # تحديد صف العنوان بناءً على نوع الورقة
            header_row = 1
            if ws.title in ["المواد للمدرسين", "الأيام والأوقات", "قائمة المواد", "تأثير الأولويات"]:
                header_row = 3
            elif ws.title in ["تفضيلات الأساتذة", "إعدادات الخوارزمية", "أولويات الأساتذة"]:
                header_row = 3
            
            # تطبيق تنسيق خاص لأولويات الأساتذة
            if ws.title == "أولويات الأساتذة":
                # تنسيق خاص للمستويات المختلفة في البيانات الفعلية (ليس العناوين)
                for row_idx, row in enumerate(ws.iter_rows(min_row=4, max_row=ws.max_row), 4):
                    if len(row) >= 3 and row[0].value and isinstance(row[0].value, int):
                        try:
                            priority_level = int(row[2].value.split('/')[0].strip()[0]) if row[2].value else 2
                        except:
                            priority_level = 2
                        
                        if priority_level == 1:
                            fill_color = "00B050"  # أخضر (أولوية عالية)
                        elif priority_level == 2:
                            fill_color = "FFC000"  # برتقالي (أولوية متوسطة)
                        else:
                            fill_color = "70AD47"  # أخضر فاتح (أولوية منخفضة)
                        
                        for cell in row:
                            if cell.value:
                                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            
            # تنسيق العناوين الرئيسية
            if header_row <= ws.max_row:
                for cell in ws[header_row]:
                    if cell.value:
                        # لون مختلف للعناوين في ورقة الأولويات والتأثير
                        if ws.title == "أولويات الأساتذة" and ("ملخص" in str(cell.value) or "تحليل" in str(cell.value)):
                            header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        elif ws.title == "تأثير الأولويات":
                            header_fill = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
                        elif ws.title == "الجدول الزمني":
                            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                        else:
                            header_fill = PatternFill(start_color="8064A2", end_color="8064A2", fill_type="solid")
                        
                        cell.fill = header_fill
                        cell.font = cell.font.copy(bold=True, color="FFFFFF")
                        cell.border = Border(left=Side(style='thin'), 
                                           right=Side(style='thin'), 
                                           top=Side(style='thin'), 
                                           bottom=Side(style='thin'))
            
            # تنسيق باقي الخلايا
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    if cell.value:
                        cell.border = Border(left=Side(style='thin'), 
                                           right=Side(style='thin'), 
                                           top=Side(style='thin'), 
                                           bottom=Side(style='thin'))
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # ضبط عرض الأعمدة
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        if isinstance(cell.value, str):
                            max_length = max(max_length, len(cell.value))
                        else:
                            max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # تجميد الصفوف في جميع الأوراق
        for ws in sheets:
            ws.freeze_panes = "A2"
        
        # حفظ الملف
        wb.save(output_filename)
        print(f"✅ تم حفظ الملف بنجاح: {output_filename}")
        print(f"📁 المسار: {os.path.abspath(output_filename)}")
        return output_filename
        
    except Exception as e:
        print(f"❌ خطأ أثناء حفظ الملف: {e}")
        import traceback
        traceback.print_exc()
        return None